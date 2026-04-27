"""
Wizard API routes — bolt onto the existing app.py Flask server.

Add to the bottom of app.py (before `if __name__ == "__main__":`)
or import via:  from wizard_routes import register_wizard_routes
                register_wizard_routes(app)
"""
from flask import Blueprint, request, jsonify
import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

wizard = Blueprint("wizard", __name__)

EXCEL_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "pipeline_output", "companies_pipeline.xlsx",
)

# ──────────────────────────────────────────────────────────────
# The canonical header row — must match run_pipeline.py + new cols
# ──────────────────────────────────────────────────────────────
HEADERS = [
    "No.", "Company Number", "Company Name", "Short Name",
    "Status", "Type", "Date of Creation", "SIC Codes", "Address",
    "Directors", "Director Nationalities",
    "DUNS Number", "DUNS Status", "DUNS Email Used",
    "Certificate Downloaded", "Certificate Path",
    "Domain", "Domain Status", "Domain Cost",
    "Assigned Email", "Account Name",
]


def _col_index(header_row, name):
    """Find column index by header name (case-insensitive)."""
    for i, h in enumerate(header_row):
        if h and str(h).strip().lower() == name.lower():
            return i
    return None


def _ensure_excel():
    """Create the pipeline Excel if it doesn't exist."""
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies Pipeline"
    ws.append(HEADERS)
    # Style headers
    hdr_fill = PatternFill(start_color="003078", end_color="003078", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)


# ──────────────────────────────────────────────────────────────
# POST /api/pipeline/add — add a company from the wizard
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/pipeline/add", methods=["POST"])
def pipeline_add():
    data = request.get_json(force=True)

    cn = (data.get("company_number") or "").strip()
    if not cn:
        return jsonify({"success": False, "error": "company_number is required"}), 400

    _ensure_excel()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Read header to find columns dynamically
    header = [cell.value for cell in ws[1]]

    # Check for duplicate
    col_cn = _col_index(header, "Company Number")
    if col_cn is not None:
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[col_cn].value and str(row[col_cn].value).strip() == cn:
                wb.close()
                return jsonify({"success": False, "error": f"Company {cn} already exists in the pipeline."}), 409

    # If "Short Name" column doesn't exist yet, add it
    col_sn = _col_index(header, "Short Name")
    if col_sn is None:
        # Insert after "Company Name"
        col_name = _col_index(header, "Company Name")
        insert_at = (col_name + 2) if col_name is not None else (len(header) + 1)
        ws.insert_cols(insert_at)
        ws.cell(row=1, column=insert_at, value="Short Name")
        # Re-read header
        header = [cell.value for cell in ws[1]]
        col_sn = _col_index(header, "Short Name")

    # Determine row number
    next_row = ws.max_row + 1
    row_num = next_row - 1  # 1-indexed company number

    def put(col_name, value):
        idx = _col_index(header, col_name)
        if idx is not None:
            ws.cell(row=next_row, column=idx + 1, value=value)

    put("No.", row_num)
    put("Company Number", cn)
    put("Company Name", data.get("company_name", ""))
    put("Short Name", data.get("short_name", ""))
    put("Status", data.get("status", ""))
    put("Type", data.get("type", ""))
    put("Date of Creation", data.get("date_of_creation", ""))
    put("SIC Codes", data.get("sic_codes", ""))
    put("Address", data.get("address", ""))
    put("Directors", data.get("directors", ""))
    put("Director Nationalities", data.get("nationalities", ""))
    put("DUNS Number", data.get("duns_number", ""))
    put("DUNS Status", "submitted" if data.get("duns_email") and not data.get("duns_number") else ("found" if data.get("duns_number") else ""))
    put("DUNS Email Used", data.get("duns_email", ""))
    put("Domain", data.get("domain", ""))
    put("Domain Status", "pending" if data.get("domain") else "")
    put("Assigned Email", data.get("email", ""))

    # Auto-width
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    wb.save(EXCEL_FILE)
    wb.close()

    return jsonify({
        "success": True,
        "row": row_num,
        "company_number": cn,
        "message": f"Added {data.get('company_name', cn)} to pipeline."
    })


# ──────────────────────────────────────────────────────────────
# GET /api/pipeline/list — list all companies in the pipeline
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/pipeline/list")
def pipeline_list():
    _ensure_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return jsonify([])

    header = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    companies = []
    for row in rows[1:]:
        entry = {}
        for i, val in enumerate(row):
            if i < len(header):
                entry[header[i]] = val
        companies.append(entry)

    return jsonify(companies)


# ──────────────────────────────────────────────────────────────
# GET /api/email-pool/status — email pool availability
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/email-pool/status")
def email_pool_status():
    try:
        from email_pool import EmailPool
        pool = EmailPool()
        status = pool.status()
        available_emails = [e for e in pool._emails if e not in pool._assignments]
        return jsonify({
            "total": status["total"],
            "used": status["used"],
            "available": status["available"],
            "available_emails": available_emails[:50],  # cap at 50
        })
    except Exception as e:
        return jsonify({"error": str(e), "available_emails": []}), 500


# ──────────────────────────────────────────────────────────────
# GET /api/domains/check — check domain availability via Namecheap
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/domains/check")
def domains_check():
    domains_str = request.args.get("domains", "").strip()
    if not domains_str:
        return jsonify({"error": "provide ?domains=example.co.uk"}), 400

    domain_list = [d.strip() for d in domains_str.split(",") if d.strip()]
    try:
        from namecheap_automation import check_domains
        results = check_domains(domain_list)
        return jsonify(results if isinstance(results, list) else [results])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# GET  /api/duns/lookup  — quick DUNS lookup
# POST /api/duns/request — submit DUNS request via stealth browser
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/duns/lookup")
def duns_lookup():
    cn = request.args.get("company_number", "").strip()
    if not cn:
        return jsonify({"error": "company_number required"}), 400
    try:
        from duns_automation import lookup_duns
        result = lookup_duns(cn)
        return jsonify(result)
    except Exception as e:
        return jsonify({"found": False, "error": str(e)})


@wizard.route("/api/duns/request", methods=["POST"])
def duns_request():
    data = request.get_json(force=True)
    cn = data.get("company_number", "").strip()
    email = data.get("email", "").strip()
    if not cn or not email:
        return jsonify({"error": "company_number and email required"}), 400
    try:
        from duns_automation import stealth_request_duns
        result = stealth_request_duns(cn, headless=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({"found": False, "error": str(e)})


# ──────────────────────────────────────────────────────────────
# Registration helper
# ──────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────
# POST /api/domains/register — actually register via Namecheap
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/domains/register", methods=["POST"])
def domains_register():
    data = request.get_json(force=True)
    domain = data.get("domain", "").strip()
    if not domain:
        return jsonify({"error": "domain is required"}), 400
    try:
        from namecheap_automation import register_domain
        result = register_domain(domain)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# POST /api/domains/find-cheapest — auto-find best domain for company
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/domains/find-cheapest", methods=["POST"])
def domains_find_cheapest():
    data = request.get_json(force=True)
    company_name = data.get("company_name", "").strip()
    if not company_name:
        return jsonify({"error": "company_name required"}), 400
    try:
        from namecheap_automation import find_cheapest_domain
        result = find_cheapest_domain(company_name)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# POST /api/email-pool/assign — assign next available email
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/email-pool/assign", methods=["POST"])
def email_pool_assign():
    data = request.get_json(force=True)
    company_number = data.get("company_number", "").strip()
    company_name = data.get("company_name", "")
    if not company_number:
        return jsonify({"error": "company_number required"}), 400
    try:
        from email_pool import EmailPool
        pool = EmailPool()
        # Check if already assigned
        existing = pool.get_assigned_email(company_number)
        if existing:
            return jsonify({"email": existing, "status": "already_assigned"})
        # Assign next
        result = pool.assign_next(company_number=company_number, company_name=company_name)
        if result:
            return jsonify({"email": result, "status": "assigned", "remaining": pool.available_count})
        return jsonify({"error": "No emails available in pool"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# POST /api/certificate/download — download incorporation certificate
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/certificate/download", methods=["POST"])
def certificate_download():
    data = request.get_json(force=True)
    company_number = data.get("company_number", "").strip()
    company_name = data.get("company_name", "")
    if not company_number:
        return jsonify({"error": "company_number required"}), 400
    try:
        import requests as http_requests
        from dotenv import load_dotenv
        load_dotenv()
        API_KEY = os.getenv("COMPANIES_HOUSE_API_KEY")
        BASE_URL = "https://api.company-information.service.gov.uk"
        # Get filing history
        resp = http_requests.get(
            f"{BASE_URL}/company/{company_number}/filing-history",
            params={"items_per_page": 100},
            auth=(API_KEY, ""),
        )
        resp.raise_for_status()
        filings = resp.json()
        # Find incorporation document
        doc_link = None
        for item in filings.get("items", []):
            cat = (item.get("category") or "").lower()
            desc = (item.get("description") or "").lower()
            ftype = (item.get("type") or "").upper()
            if cat == "incorporation" or ftype == "NEWINC" or "incorporat" in desc:
                links = item.get("links", {})
                doc_link = links.get("document_metadata")
                break
        if not doc_link:
            return jsonify({"status": "not_found", "message": "No incorporation filing found"})
        # Download PDF
        meta_resp = http_requests.get(f"https://api.company-information.service.gov.uk{doc_link}",
                                      auth=(API_KEY, ""))
        meta_resp.raise_for_status()
        meta = meta_resp.json()
        pdf_url = meta.get("links", {}).get("document")
        if not pdf_url:
            return jsonify({"status": "no_pdf", "message": "Document metadata found but no PDF link"})
        # Download the PDF
        pdf_resp = http_requests.get(pdf_url, auth=(API_KEY, ""),
                                     headers={"Accept": "application/pdf"})
        if pdf_resp.status_code == 200:
            safe_name = company_name.replace("/", "-").replace("\\", "-").strip() or company_number
            cert_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "certificates", safe_name)
            os.makedirs(cert_dir, exist_ok=True)
            cert_path = os.path.join(cert_dir, f"Certificate_{company_number}.pdf")
            with open(cert_path, "wb") as f:
                f.write(pdf_resp.content)
            return jsonify({"status": "downloaded", "path": cert_path, "size": len(pdf_resp.content)})
        return jsonify({"status": "download_failed", "http_status": pdf_resp.status_code})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ──────────────────────────────────────────────────────────────
# POST /api/domains/setup-email-forwarding
# ──────────────────────────────────────────────────────────────
@wizard.route("/api/domains/setup-email-forwarding", methods=["POST"])
def domains_email_forwarding():
    data = request.get_json(force=True)
    domain = data.get("domain", "").strip()
    forward_to = data.get("forward_to", "").strip()
    if not domain or not forward_to:
        return jsonify({"error": "domain and forward_to required"}), 400
    try:
        from namecheap_automation import set_email_forwarding
        result = set_email_forwarding(domain, "support", forward_to)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def register_wizard_routes(app):
    """Call this from app.py to mount all wizard endpoints."""
    app.register_blueprint(wizard)
